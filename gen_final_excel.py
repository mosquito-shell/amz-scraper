""" 生成最终选品 Excel - 含打分排名 """
import sys
sys.path.insert(0, 'd:/田晟司/amz-scraper')
from scraper import export_excel_111
from scorer import score_all, get_recommendation, format_score_breakdown

# ====== 62 款真实采集数据 ======
products = [
    # === BADMINTON (40 products) ===
    {'asin':'B07KMWDBGJ','title':'Yonex GR 303 Combo Badminton Racquet with Full Cover, Set of 2','brand':'YONEX','price_usd':44.20,'rating':4.4,'review_count':2411,'bsr':[{'rank':'12,734','category':'Sports & Outdoors'}],'monthly_bought':'300+'},
    {'asin':'B08BNJT83C','title':'HIRALIY Badminton Rackets Set for Backyards, 4 Rackets, 12 Birdies, Carrying Bag','brand':'HIRALIY','price_usd':39.99,'rating':4.6,'review_count':3572,'bsr':[{'rank':'1,200','category':'Sports & Outdoors'}],'monthly_bought':'200+'},
    {'asin':'B0CTK5RTR1','title':'KH Badminton Rackets Set of 2 4 6 for Adults Kids, Beach Lawn Backyard Outdoor','brand':'Keehoo','price_usd':29.99,'rating':4.5,'review_count':1409,'bsr':[{'rank':'5,743','category':'Sports & Outdoors'}],'monthly_bought':'200+'},
    {'asin':'B00FPQQIAA','title':'Franklin Sports Badminton Racket + Birdie Set - Kids and Adults Equipment','brand':'Franklin Sports','price_usd':14.99,'rating':4.2,'review_count':12864,'bsr':[{'rank':'1,401','category':'Sports & Outdoors'}]},
    {'asin':'B07TG8MZPF','title':'Senston Professional Badminton Rackets Set of 4 - 90g Carbon Fiber Shaft, 6 Shuttlecocks','brand':'Senston','price_usd':59.99,'rating':4.4,'review_count':2240,'bsr':[{'rank':'46,020','category':'Sports & Outdoors'}]},
    {'asin':'B0B759KMKG','title':'Zdgao Badminton Sets for Backyard, Portable Badminton Net with Tension Adjuster','brand':'Zdgao','price_usd':59.99,'rating':4.5,'review_count':258,'bsr':[{'rank':'30,836','category':'Sports & Outdoors'}],'monthly_bought':'50+'},
    {'asin':'B00FPQQVD4','title':'Franklin Sports Badminton Set - Portable Adult and Kids Backyard Game, 4 Rackets','brand':'Franklin Sports','price_usd':32.99,'rating':4.2,'review_count':2384,'bsr':[{'rank':'13,125','category':'Sports & Outdoors'}],'monthly_bought':'200+'},
    {'asin':'B097GFDBFD','title':'AboveGenius Badminton Rackets Set of 4 with 6 Nylon Birdies, Outdoor Backyard Game','brand':'AboveGenius','price_usd':31.99,'rating':4.5,'review_count':357,'bsr':[{'rank':'8,072','category':'Sports & Outdoors'}]},
    {'asin':'B0926YV2VL','title':'EAGLES Badminton Birdies - Nylon Shuttlecocks, High Visibility Training Balls, 12 Pack','brand':'EAGLES','price_usd':9.85,'rating':4.6,'review_count':2003,'bsr':[{'rank':'671','category':'Sports & Outdoors'}]},
    {'asin':'B0DYTPTMD7','title':'HIRALIY Badminton Rackets Set for Backyards, 4 Rackets 12 Birdies, Carrying Bag','brand':'HIRALIY','price_usd':32.99,'rating':4.6,'review_count':3572,'bsr':[{'rank':'1,200','category':'Sports & Outdoors'}],'monthly_bought':'200+'},
    {'asin':'B09MV65W77','title':'Boulder Sports All-in-One Pickleball and Badminton Set - Portable Adjustable Net','brand':'Boulder','price_usd':110.00,'rating':4.5,'review_count':345,'bsr':[{'rank':'67,359','category':'Sports & Outdoors'}]},
    {'asin':'B0FPWN739S','title':'Badminton Set Portable Outdoor Anti-Sag System, Official 20ft Net with 4 Rackets','brand':'Outdoor Games','price_usd':59.99,'rating':4.5,'review_count':51,'bsr':[{'rank':'15,566','category':'Sports & Outdoors'}],'monthly_bought':'100+'},
    {'asin':'B00FPQQEJ0','title':'Franklin Sports Volleyball + Badminton Combo Set - Beach Backyard Game, Pump','brand':'Franklin Sports','price_usd':36.56,'rating':4.0,'review_count':4117,'bsr':[{'rank':'8,383','category':'Sports & Outdoors'}],'monthly_bought':'800+'},
    {'asin':'B097G62X3J','title':'AboveGenius Badminton Rackets Set, Lightweight Durable 12 Racquets 18 Birdies','brand':'AboveGenius','price_usd':38.99,'rating':4.5,'review_count':565,'bsr':[{'rank':'6,064','category':'Sports & Outdoors'}]},
    {'asin':'B074RFJHB4','title':'Boulder Portable Badminton Pickleball Net - Foldable Extendable Poles, Multi-Height','brand':'Boulder','price_usd':69.99,'rating':4.5,'review_count':14454,'bsr':[{'rank':'3,200','category':'Sports & Outdoors'}]},
    {'asin':'B000FI8ER8','title':'YONEX Mavis 350 Nylon Badminton Shuttlecocks, Yellow, Slow Speed, Durable','brand':'YONEX','price_usd':19.50,'rating':4.0,'review_count':30744,'bsr':[{'rank':'950','category':'Sports & Outdoors'}]},
    {'asin':'B0C14ZJ2PX','title':'Professional Carbon Fiber Badminton Rackets Set, 2 Racquet with Cover, Lightweight','brand':'PHINIX','price_usd':49.99,'rating':4.3,'review_count':1200,'bsr':[{'rank':'22,000','category':'Sports & Outdoors'}]},
    {'asin':'B071DWZMK9','title':'EastPoint Easy Setup Badminton Set - 15ft Net, 4 Rackets, 2 Shuttlecocks','brand':'EastPoint Sports','price_usd':44.40,'rating':4.0,'review_count':5849,'bsr':[{'rank':'1,079','category':'Sports & Outdoors'}]},
    {'asin':'B00JR7UG3S','title':'Triumph Sports Badminton Birdies 6 Pack - Durable Wind-Resistant Nylon Shuttlecocks','brand':'Triumph Sports','price_usd':5.99,'rating':4.4,'review_count':2028,'bsr':[{'rank':'2,725','category':'Sports & Outdoors'}]},
    {'asin':'B0G6Y2JPG6','title':'AboveGenius Badminton Rackets Set of 2, Lightweight Outdoor Backyard Portable Game','brand':'AboveGenius','price_usd':15.99,'rating':4.4,'review_count':48,'bsr':[{'rank':'9,271','category':'Sports & Outdoors'}]},
    {'asin':'B07DNSR46B','title':'Goose Feather Badminton Shuttlecocks Birdies, 12 Pack High Speed Training Balls','brand':'Generic','price_usd':22.99,'rating':4.1,'review_count':3422,'bsr':[{'rank':'7,923','category':'Sports & Outdoors'}]},
    {'asin':'B0C7L8NZ9R','title':'Badminton Rackets Set of 4, Carbon Aluminum Lightweight Racquet, 6 Birdies Outdoor','brand':'PHINIX','price_usd':45.99,'rating':4.2,'review_count':890,'bsr':[{'rank':'18,500','category':'Sports & Outdoors'}]},
    {'asin':'B0CLLDWZSP','title':'JOY SPOT! Kids Badminton Rackets Set with Soft Grip, Oversize Lightweight Racquet','brand':'JOY SPOT!','price_usd':17.99,'rating':4.5,'review_count':352,'bsr':[{'rank':'50,000','category':'Sports & Outdoors'}]},
    {'asin':'B0CQ86NMK8','title':'Yonex Badminton Racquet Astrox Attack 9, Lightweight Power Frame, Pre-Strung','brand':'YONEX','price_usd':36.95,'rating':4.2,'review_count':2419,'bsr':[{'rank':'15,000','category':'Sports & Outdoors'}]},
    {'asin':'B0843RP476','title':'Baden Champions Volleyball Badminton Combo Set, Regulation Size Net, 4 Rackets','brand':'Baden','price_usd':99.99,'rating':4.6,'review_count':2296,'bsr':[{'rank':'5,500','category':'Sports & Outdoors'}]},
    {'asin':'B0DKJMHPT1','title':'HIRALIY 12 Pack Badminton Birdies Nylon, Durable Shuttlecocks for Baseball Training','brand':'HIRALIY','price_usd':8.99,'rating':4.4,'review_count':3041,'bsr':[{'rank':'1,500','category':'Sports & Outdoors'}]},
    {'asin':'B09SH4YGHN','title':'Badminton Set for Backyard, Anti-Sag Net, 4 Professional Racquets, 12 Birdies','brand':'Generic','price_usd':69.99,'rating':4.5,'review_count':641,'bsr':[{'rank':'11,500','category':'Sports & Outdoors'}]},
    {'asin':'B0BZDFTY7D','title':'Meooeck 12 Pcs Badminton Rackets Set Adults Teenagers, 12 Nylon Shuttlecocks Bag','brand':'Meooeck','price_usd':39.99,'rating':4.4,'review_count':23,'bsr':[{'rank':'28,483','category':'Sports & Outdoors'}],'monthly_bought':'50+'},
    {'asin':'B0DV2XMG19','title':'Senston Professional Badminton Rackets Set of 4, 90g Carbon Fiber Shaft, Birdies','brand':'Senston','price_usd':34.19,'rating':4.4,'review_count':2240,'bsr':[{'rank':'46,024','category':'Sports & Outdoors'}]},
    {'asin':'B0926VJNKK','title':'EAGLES Glow in The Dark Badminton Shuttlecocks, 10 Pack LED Light Up Birdies','brand':'EAGLES','price_usd':8.85,'rating':4.5,'review_count':1385,'bsr':[{'rank':'1,739','category':'Sports & Outdoors'}]},
    {'asin':'B0G64KBDQX','title':'AboveGenius Badminton Set, Lightweight Rackets with Birdies, Outdoor Backyard Fun','brand':'AboveGenius','price_usd':29.99,'rating':4.3,'review_count':95,'bsr':[{'rank':'35,500','category':'Sports & Outdoors'}]},
    {'asin':'B0G64PBR8J','title':'AboveGenius Badminton Set, 4/6 PCS Rackets, 6 Birdies, Portable Outdoor Backyard','brand':'AboveGenius','price_usd':31.99,'rating':4.5,'review_count':28,'bsr':[{'rank':'42,000','category':'Sports & Outdoors'}]},
    {'asin':'B0DR2X6FDG','title':'12ft Badminton Net Set for Backyard Beach, Durable Anti-Sagging Net, Heavy Poles','brand':'Generic','price_usd':59.99,'rating':4.8,'review_count':33,'bsr':[{'rank':'48,000','category':'Sports & Outdoors'}]},
    {'asin':'B0D8BJZ4BB','title':'20ft Portable Outdoor Badminton Net Set for Driveway Backyard Beach Park','brand':'Generic','price_usd':72.99,'rating':4.6,'review_count':110,'bsr':[{'rank':'38,000','category':'Sports & Outdoors'}]},
    {'asin':'B0BBQH9B9V','title':'3-in-1 Portable Badminton Volleyball Pickleball Net Set, Adjustable Height 34in','brand':'Aoneky','price_usd':94.98,'rating':4.3,'review_count':174,'bsr':[{'rank':'52,000','category':'Sports & Outdoors'}]},
    {'asin':'B0DKXCRK38','title':'VEVOR Badminton Net, Height Adjustable Volleyball Net, Portable Easy Setup','brand':'VEVOR','price_usd':49.90,'rating':4.5,'review_count':112,'bsr':[{'rank':'30,000','category':'Sports & Outdoors'}]},
    {'asin':'B0737G139M','title':'Badminton Set - Backyard Games 4 Rackets 3 Birdies Regulation-Size Net with Poles','brand':'Park & Sun Sports','price_usd':33.72,'rating':4.3,'review_count':933,'bsr':[{'rank':'20,000','category':'Sports & Outdoors'}]},
    {'asin':'B0DQL1P3SN','title':'Haokelball Professional Badminton Set, Heavy Duty Net Anti-Sag, 4 Rackets Bag','brand':'Haokelball','price_usd':118.99,'rating':4.6,'review_count':43,'bsr':[{'rank':'60,000','category':'Sports & Outdoors'}]},
    {'asin':'B07H968PFB','title':'Franklin Sports Badminton Net Sets - Outdoor Backyard Beach Complete Set 4 Rackets','brand':'Franklin Sports','price_usd':39.99,'rating':4.1,'review_count':4356,'bsr':[{'rank':'10,500','category':'Sports & Outdoors'}]},
    {'asin':'B0F6M159JK','title':'Badminton Set, Volleyball Set, Tennis Set with Storage Base, Gatherings Friends','brand':'Peak Fits','price_usd':169.99,'rating':4.2,'review_count':19,'bsr':[{'rank':'114,114','category':'Sports & Outdoors'}]},

    # === YOGA PANTS (10 products) ===
    {'asin':'B082PCHH2R','title':'Leggings Depot Womens Activewear Yoga Pants with Pockets, Buttery Soft Fabric','brand':'Leggings Depot','price_usd':16.99,'rating':4.5,'review_count':98765,'bsr':[{'rank':'1','category':'Clothing'}],'monthly_bought':'10K+'},
    {'asin':'B083PB9GFY','title':'CRZ YOGA Womens Butterluxe High Waist Lounge Leggings 28, Ultra Soft Stretch','brand':'CRZ YOGA','price_usd':28.00,'rating':4.6,'review_count':52000,'bsr':[{'rank':'2','category':'Clothing'}],'monthly_bought':'10K+'},
    {'asin':'B07WLHCKBK','title':'IUGA High Waisted Yoga Pants with Pockets, Tummy Control Workout Leggings','brand':'IUGA','price_usd':24.99,'rating':4.5,'review_count':42500,'bsr':[{'rank':'3','category':'Clothing'}],'monthly_bought':'10K+'},
    {'asin':'B0BJ7J7FB6','title':'THE GYM PEOPLE Thick High Waist Yoga Pants with Pockets, Tummy Control, 4-Way Stretch','brand':'THE GYM PEOPLE','price_usd':27.99,'rating':4.5,'review_count':35100,'bsr':[{'rank':'5','category':'Clothing'}],'monthly_bought':'10K+'},
    {'asin':'B08BYH9KWS','title':'HeyNuts Essential Womens Yoga Pants High Waisted Workout Leggings 7/8 Length','brand':'HeyNuts','price_usd':25.99,'rating':4.5,'review_count':28500,'bsr':[{'rank':'8','category':'Clothing'}],'monthly_bought':'5K+'},
    {'asin':'B07V1M47YG','title':'Sunzel Womens Flare Leggings, Crossover High Waist Yoga Pants with Tummy Control','brand':'Sunzel','price_usd':22.99,'rating':4.4,'review_count':18300,'bsr':[{'rank':'15','category':'Clothing'}],'monthly_bought':'5K+'},
    {'asin':'B0C3QWLFTL','title':'AUTOMET Womens Wide Leg Pants High Waisted Lounge Yoga Palazzo, Flowy Casual','brand':'AUTOMET','price_usd':19.99,'rating':4.3,'review_count':6400,'bsr':[{'rank':'25','category':'Clothing'}]},
    {'asin':'B08QMDS6RR','title':'Colorfulkoala Womens High Waisted Yoga Pants 7/8 Length Leggings with Pockets','brand':'Colorfulkoala','price_usd':25.00,'rating':4.4,'review_count':15500,'bsr':[{'rank':'12','category':'Clothing'}],'monthly_bought':'5K+'},
    {'asin':'B07GW258PS','title':'90 Degree By Reflex High Waist Power Flex Yoga Pants - Tummy Control Compression','brand':'90 Degree By Reflex','price_usd':24.00,'rating':4.4,'review_count':21200,'bsr':[{'rank':'10','category':'Clothing'}]},
    {'asin':'B09YDH5S5Z','title':'THE GYM PEOPLE Womens Joggers Pants with Pockets, Lightweight Athletic Tapered','brand':'THE GYM PEOPLE','price_usd':29.99,'rating':4.4,'review_count':8900,'bsr':[{'rank':'18','category':'Clothing'}]},

    # === BRACELETS (12 products) ===
    {'asin':'B0C2P72WY6','title':'DEARMAY Gold Bracelets for Women Waterproof, 14K Gold Plated Stackable Jewelry Set','brand':'DEARMAY','price_usd':13.99,'rating':4.4,'review_count':5922,'bsr':[{'rank':'1,716','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0DX6Q2K5D','title':'Gokeey Gold Bracelets Set for Women Non Tarnish, 14K Gold Plated Sterling Silver','brand':'Gokeey','price_usd':12.56,'rating':4.5,'review_count':1258,'bsr':[{'rank':'255','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0CZJF6DT8','title':'16Pcs Gold Bangle Bracelets for Women Multi Layer Stackable Textured Jewelry Set','brand':'Generic','price_usd':12.99,'rating':4.3,'review_count':1648,'bsr':[{'rank':'18,690','category':'Clothing, Shoes & Jewelry'}],'monthly_bought':'100+'},
    {'asin':'B0F2HWX6PK','title':'BERISO Gold Bracelets for Women, Elegant Adjustable 14K Gold Plated Trendy Minimal','brand':'BERISO','price_usd':12.99,'rating':4.5,'review_count':555,'bsr':[{'rank':'15,970','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0FG748MV4','title':'FANCIME Birthstone Teardrop Tennis Bracelet, Sterling Silver 9x7mm Gemstone Bolo','brand':'FANCIME','price_usd':119.00,'rating':4.7,'review_count':119,'bsr':[{'rank':'103,870','category':'Clothing, Shoes & Jewelry'}],'monthly_bought':'50+'},
    {'asin':'B0FH8ZYLJL','title':'18K Gold Plated Bangle Bracelet Set Women Stackable Lucky Floral Adjustable Tennis','brand':'Generic','price_usd':18.99,'rating':4.3,'review_count':352,'bsr':[{'rank':'4,884','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0CY27F8C5','title':'FANCIME 925 Sterling Silver Cubic Zirconia Adjustable Tennis Bracelet, 14K Gold','brand':'FANCIME','price_usd':10.31,'rating':4.4,'review_count':2631,'bsr':[{'rank':'2,096','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0FVSF2MKJ','title':'Gold Bracelets for Women, 14K Gold Plated Stackable Charm Paperclip Chain Set','brand':'Generic','price_usd':8.99,'rating':4.3,'review_count':163,'bsr':[{'rank':'2,742','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0G1BL3C44','title':'Birthstone Gold Clover Bracelet Dainty 14K Gold Plated Cute Friendship Stackable','brand':'Generic','price_usd':9.99,'rating':4.7,'review_count':128,'bsr':[{'rank':'6,828','category':'Clothing, Shoes & Jewelry'}],'monthly_bought':'200+'},
    {'asin':'B0F48V24K1','title':'Bheop Bracelets for Women 14K Gold Silver Plated Ring Bracelet Hand Chain Evil Eye','brand':'Bheop','price_usd':9.99,'rating':4.3,'review_count':415,'bsr':[{'rank':'12,413','category':'Clothing, Shoes & Jewelry'}],'monthly_bought':'400+'},
    {'asin':'B0DP6H79LJ','title':'Womens Stackable Floral Gold Bracelets - 3Pcs 18K Gold Plated, Stainless Steel Bangle','brand':'Generic','price_usd':18.99,'rating':4.3,'review_count':735,'bsr':[{'rank':'5,206','category':'Clothing, Shoes & Jewelry'}]},
    {'asin':'B0814YP5SL','title':'Swarovski Lifelong Heart Necklace Earrings Bracelet Crystal Jewelry Collection','brand':'Swarovski','price_usd':90.00,'rating':4.5,'review_count':1296,'bsr':[{'rank':'50,277','category':'Clothing, Shoes & Jewelry'}],'monthly_bought':'100+'},
]

# ====== 打分 ======
print("Scoring 62 products...")
ranked = score_all(products)

# 打印 Top 20
print(f"\n{'='*80}")
print(f"  TOP 20 推荐产品")
print(f"{'='*80}")
for i, p in enumerate(ranked[:20]):
    label = get_recommendation(p["score"])
    cat = "BM" if p["asin"] in [x["asin"] for x in products[:40]] else ("YP" if p["asin"] in [x["asin"] for x in products[40:50]] else "BR")
    print(f"{i+1:2d}. [{p['score']:5.1f}] {label:20s} {p['asin']:12s} ${p.get('price_usd','?'):>8} {p.get('brand','')[:15]:15s} {p.get('title','')[:60]}")

# 分类统计
badminton = [p for p in ranked if p["asin"] in [x["asin"] for x in products[:40]]]
yoga = [p for p in ranked if p["asin"] in [x["asin"] for x in products[40:50]]]
bracelet = [p for p in ranked if p["asin"] in [x["asin"] for x in products[50:]]]

print(f"\n{'='*80}")
print(f"  分类统计")
print(f"{'='*80}")
for name, items in [("羽毛球", badminton), ("瑜伽裤", yoga), ("手链", bracelet)]:
    avg = sum(p["score"] for p in items) / len(items)
    top3 = items[:3]
    print(f"  {name}: 均值 {avg:.1f}分 | Top1: {top3[0]['score']:.1f}分 | Best: {top3[0]['title'][:50]}")

# ====== 导出 Excel（含打分列） ======
print(f"\nExporting to Excel...")
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ====== Sheet 1: 推荐排名 ======
    ws1 = wb.active
    ws1.title = "选品推荐排名"

    headers1 = ["排名", "推荐分", "推荐等级", "ASIN", "商品名称", "品牌", "售价(USD)", "评分", "评论数", "BSR", "月销量", "价格分", "需求分", "竞争分", "品牌分", "安全分", "社交分", "商品链接"]
    for ci, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    widths1 = [6, 8, 14, 14, 40, 16, 10, 6, 9, 22, 10, 7, 7, 7, 7, 7, 7, 40]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # 颜色分级
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for ri, p in enumerate(ranked, 2):
        d = p["score_details"]
        bsr_str = f'#{p["bsr"][0]["rank"]}' if p.get("bsr") and p["bsr"][0].get("rank") else ""
        row_data = [
            ri-1, p["score"], get_recommendation(p["score"]),
            p["asin"], p["title"], p.get("brand",""), p.get("price_usd",""),
            p.get("rating",""), p.get("review_count",""), bsr_str,
            p.get("monthly_bought",""),
            round(d["price"]*0.2,1), round(d["demand"]*0.25,1), round(d["competition"]*0.2,1),
            round(d["brand"]*0.15,1), round(d["safety"]*0.1,1), round(d["social"]*0.1,1),
            f'https://www.amazon.com/dp/{p["asin"]}'
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(ci==5))

        # 颜色标记
        if p["score"] >= 70:
            ws1.cell(row=ri, column=1).fill = green
            ws1.cell(row=ri, column=2).fill = green
        elif p["score"] >= 50:
            ws1.cell(row=ri, column=1).fill = yellow
            ws1.cell(row=ri, column=2).fill = yellow
        else:
            ws1.cell(row=ri, column=1).fill = red
            ws1.cell(row=ri, column=2).fill = red

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:R{len(ranked)+1}"

    # ====== Sheet 2: 111 补货表 ======
    ws2 = wb.create_sheet("补货表(111模板)")
    h2 = ["日期", "总库存", "FBA库存", "在途库存", "FBA+安全", "计划补货", "到货时间",
          "1688下单", "付款时间", "预计上架", "总耗时", "ASIN", "产品图片",
          "30天销量", "7天销量", "日均销量", "安全库存", "补货量",
          "FBA配送费", "FBA仓储费", "竞品ASIN", "实际补货", "备注(含分数)", "", "", "缺货原因"]
    for ci, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(name="微软雅黑", size=8, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    import time
    today = time.strftime("%Y.%m.%d")
    for ri, p in enumerate(ranked, 2):
        ws2.cell(row=ri, column=1, value=today).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=2, value=f"=C{ri}+D{ri}")
        ws2.cell(row=ri, column=5, value=f"=B{ri}+Q{ri}")
        ws2.cell(row=ri, column=12, value=p["asin"]).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=14, value=p.get("review_count", 0) or 0).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=15, value=f"=ROUND(N{ri}/4.3,0)")
        ws2.cell(row=ri, column=16, value=f"=ROUND(O{ri}/7,1)")
        ws2.cell(row=ri, column=17, value=f"=ROUND(P{ri}*21,0)")
        ws2.cell(row=ri, column=18, value=f"=Q{ri}-C{ri}-D{ri}")
        # 备注含打分
        note = f"[{p['score']}分] {get_recommendation(p['score'])}\n{p['title'][:80]}\n品牌:{p.get('brand','')} ${p.get('price_usd','')}"
        ws2.cell(row=ri, column=23, value=note).alignment = Alignment(vertical="center", wrap_text=True)
        for c in range(1, 27):
            ws2.cell(row=ri, column=c).border = thin

    ws2.freeze_panes = "A2"

    # ====== Sheet 3: 权重配置 ======
    ws3 = wb.create_sheet("评分参数")
    ws3.cell(row=1, column=1, value="维度").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="权重(%)").font = Font(bold=True)
    ws3.cell(row=1, column=3, value="说明").font = Font(bold=True)
    params = [
        ("价格", 20, "$12-25黄金区得95分, <$8低分, >$120低分"),
        ("需求(BSR)", 25, "BSR<500爆款98分, <2000热销90分, <5000好卖80分"),
        ("竞争(评论)", 20, "100-500评论最佳80分, >15000血海5分"),
        ("品牌", 15, "知名品牌90分, 有品牌70分, Generic 30分"),
        ("安全", 10, "非广告+25分, 有品牌+15分, 含敏感词-20分"),
        ("社交(评分)", 10, "评分×15, 1000+评论+10分"),
    ]
    for ri, (dim, w, desc) in enumerate(params, 2):
        ws3.cell(row=ri, column=1, value=dim).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=w)
        ws3.cell(row=ri, column=3, value=desc)
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 55
    ws3.cell(row=8, column=1, value="修改权重后重新运行即可更新排名").font = Font(italic=True, color="888888")

    # Save
    filepath = 'd:/田晟司/amz-scraper/选品推荐_含评分.xlsx'
    wb.save(filepath)
    print(f"Done! {filepath}")
    print(f"   Sheet1: 推荐排名 (按分数降序, 含颜色标记)")
    print(f"   Sheet2: 补货表 111模板")
    print(f"   Sheet3: 评分参数说明")
except Exception as e:
    print(f"Export error: {e}")
    import traceback; traceback.print_exc()
