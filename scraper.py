"""
亚马逊商品数据批量抓取脚本
通过 CF Worker 代理抓取搜索页 + 商品详情 + 图片下载
输出格式: 领翼-模版 / 111.xlsx 兼容

用法:
    python scraper.py search "anker charger" --pages 3 --export result.xlsx --images
    python scraper.py product B0C8HHV9DK --images
    python scraper.py bestsellers --count 50 --export best.xlsx --images
"""
import httpx
import re
import json
import sys
import time
import argparse
import os
import urllib3
from urllib.parse import quote, urljoin, urlparse

urllib3.disable_warnings()

# ====== HS 编码推荐引擎 (与插件同步) ======
_HS_DB = [
    ("racket","badminton","tennis","paddle","shuttlecock","9506.59","运动球拍/羽毛球",4.7),
    ("ball","soccer","basketball","football","volleyball","9506.62","可充气球类",3.0),
    ("yoga","legging","tights","pant","stretch","6112.49","瑜伽裤/紧身裤(化纤)",28.4),
    ("cotton","shirt","jean","denim","6204.62","棉质长裤",16.6),
    ("jacket","coat","parka","hoodie","sweatshirt","6201.93","化纤外套/夹克",27.7),
    ("bra","underwear","brief","panty","boxer","6212.10","内衣/文胸",16.9),
    ("swim","bikini","swimwear","trunk","6211.11","泳装(男)",7.6),
    ("sock","stocking","hosiery","6115.96","袜类(化纤)",18.8),
    ("hat","cap","baseball","beanie","visor","6505.00","帽类",7.5),
    ("bag","backpack","handbag","purse","tote","4202.22","手提包(化纤)",17.6),
    ("wallet","card","holder","4202.31","钱包(皮革)",8.0),
    ("shoe","sneaker","boot","sandal","slipper","6404.19","运动鞋(纺织面)",20.0),
    ("watch","wristwatch","timepiece","9102.11","电子手表",4.4),
    ("ring","pendant","necklace","gold","7113.19","贵金属首饰",5.5),
    ("bracelet","bangle","cuff","anklet","7117.90","仿首饰/手链",0.0),
    ("necklace","pendant","choker","locket","7117.90","仿首饰/项链",0.0),
    ("charger","usb","adapter","power","8504.40","充电器/电源适配器",2.5),
    ("cable","cord","wire","lightning","8544.42","USB数据线",2.6),
    ("speaker","bluetooth","audio","sound","8518.22","蓝牙音箱",4.9),
    ("headphone","earphone","earbud","headset","8518.30","耳机",4.9),
    ("keyboard","mouse","mechanical","gaming","8471.60","键盘/鼠标",0.0),
    ("monitor","screen","display","lcd","8528.52","显示器",0.0),
    ("battery","power","bank","recharge","8507.60","锂电池/充电宝",3.4),
    ("lamp","light","bulb","led","flashlight","9405.40","LED灯具",5.3),
    ("mat","carpet","rug","floor","5703.30","地毯/地垫(化纤)",6.0),
    ("towel","wash","beach","microfiber","6302.60","毛巾(棉)",9.1),
    ("blanket","throw","fleece","quilt","6301.40","毯子(化纤)",8.5),
    ("pillow","cushion","bolster","9404.90","枕头/靠垫",6.0),
    ("bottle","tumbler","vacuum","stainless","9617.00","保温杯(不锈钢)",8.0),
    ("cup","mug","ceramic","coffee","6912.00","陶瓷杯/餐具",6.0),
    ("toy","figure","doll","action","plush","9503.00","玩具/玩偶",0.0),
    ("dumbbell","weight","kettlebell","barbell","9506.91","哑铃/健身器械",3.0),
    ("resistance","band","tube","loop","9506.91","弹力带/健身配件",3.0),
    ("goggle","glasses","swim","dive","snorkel","9004.90","护目镜/游泳镜",2.5),
    ("camera","dashcam","camcorder","gopro","8525.89","相机/摄像机",0.0),
    ("sunglasses","eyewear","shades","9004.10","太阳镜",2.0),
    ("bike","cycling","bicycle","helmet","8712.00","自行车",11.0),
    ("skate","skateboard","longboard","9506.70","滑板",0.0),
    ("glove","boxing","mma","martial","9511.20","拳击/格斗手套",3.8),
    ("phone","case","cover","iphone","4202.32","手机壳(化纤)",17.6),
    ("screen","protector","tempered","film","3920.69","屏幕保护膜(塑料)",5.8),
    ("kitchen","utensil","spatula","whisk","3924.10","厨房用具(塑料)",6.5),
    ("scissors","shear","cutter","craft","8213.00","剪刀",2.7),
    ("knife","pocket","folding","blade","8211.92","刀具(固定刀片)",3.9),
    ("pet","dog","cat","leash","collar","4201.00","宠物用品",0.8),
    ("massage","gun","massager","fascia","9019.10","按摩器",1.4),
    ("vitamin","supplement","pill","capsule","2106.90","膳食补充剂",6.4),
    ("makeup","cosmetic","lipstick","mascara","3304.20","化妆品",0.0),
    ("jump","rope","skipping","speed","9506.91","跳绳",3.0),
    ("tent","camping","sleeping","bag","hiking","6306.22","帐篷(化纤)",8.8),
]

def _hs_guess(title, brand=""):
    if not title: return "6117.90","纺织品杂项",7.5
    t = title.lower()
    best, best_score = None, 0
    words = [w for w in t.replace("/"," ").replace("-"," ").split() if len(w) > 2]
    words.append(t[:40])
    for row in _HS_DB:
        score = 0
        for i in range(5):
            if row[i] and len(row[i]) > 1 and row[i] in t: score += 15
        for w in words:
            for j in range(5):
                if row[j] == w: score += 25
        if row[5] in t: score += 30
        if score > best_score: best_score, best = score, row
    if best and best_score >= 15: return best[5], best[6], best[7]
    return "6117.90", "纺织品杂项", 7.5

PROXY_BASE = "https://proxy.tsscjn.top"
PROXY_POOL = [
    "https://proxy.tsscjn.top",
    "https://amz-proxy-2.3203916089.workers.dev",
    "https://amz-proxy-3.3203916089.workers.dev",
    "https://amz-proxy-4.3203916089.workers.dev",
]

# 加载 IP 池 (proxyIP_cache.json)
_IP_POOL = []
def _load_ip_pool():
    """从 proxyIP_cache.json 加载可用 IP"""
    global _IP_POOL
    for path in ["proxyIP_cache.json", os.path.join(os.path.dirname(__file__), "proxyIP_cache.json")]:
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    cache = json.load(f)
                _IP_POOL = [(ip["ip"], ip.get("port", 443)) for ip in cache.get("ips", []) if ip.get("ip")]
                if _IP_POOL:
                    print(f"  [IP池] 加载 {len(_IP_POOL)} 个 IP 作为备用代理")
                return
        except: pass
    _IP_POOL = []

_load_ip_pool()

# ====== IP 池主动轮换: 记录已用 IP, 不重复使用直到池耗尽 ======
_IP_USED = set()  # 本轮已用过的 IP
_IP_CURSOR = 0   # 轮换游标

def _get_next_ip():
    """返回下一个未用过的代理IP (HTTP代理格式). 池耗尽后重置."""
    global _IP_CURSOR, _IP_USED
    if not _IP_POOL:
        return None
    # Filter valid proxy IPs only (port 80/3128/8080 — 开放代理端口)
    usable = [(ip,p) for ip,p in _IP_POOL if p in (80,3128,8080,1080,8888,9090,8000,8118)]
    if not usable:
        usable = _IP_POOL  # fallback to all
    # Remove already-used IPs
    fresh = [(ip,p) for ip,p in usable if ip not in _IP_USED]
    if not fresh:
        _IP_USED.clear()
        fresh = usable
    ip, port = fresh[_IP_CURSOR % len(fresh)]
    _IP_CURSOR += 1
    _IP_USED.add(ip)
    return f"http://{ip}:{port}"

# ====== 第1层: UA 池 (20个真实 Chrome UA, 随机切换) ======
import random as _random
UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (X11; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; CrOS x86_64 14541.0.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 OPR/111.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 OPR/111.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]
LANG_POOL = ["en-US,en;q=0.9", "en-GB,en;q=0.8", "en-US,en;q=0.8", "en;q=0.9", "en-US,en;q=0.7"]

def random_headers():
    """每次请求生成随机浏览器头"""
    return {
        "User-Agent": _random.choice(UA_POOL),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": _random.choice(LANG_POOL),
        "Accept-Encoding": "gzip, deflate, br",
        "Cache-Control": "no-cache",
        "Sec-Ch-Ua": '"Google Chrome";v="125", "Chromium";v="125", "Not.A/Brand";v="24"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
    }

# ====== 第2层: IP 冷却池 ======
COOLDOWN = {}  # {proxy_url: cooldown_until_timestamp}

def get_active_proxy(attempt=0):
    """返回未被冷却的 Worker 或 IP 池代理"""
    now = time.time()
    # 清理过期冷却
    for key in list(COOLDOWN.keys()):
        if COOLDOWN[key] < now:
            del COOLDOWN[key]
    # 找可用 Worker
    available = [p for p in PROXY_POOL if p not in COOLDOWN]
    if available:
        return available[attempt % len(available)]
    # Workers 全封 → 使用 IP 池直连
    if _IP_POOL:
        idx = attempt % len(_IP_POOL)
        ip, port = _IP_POOL[idx]
        proxy_url = f"http://{ip}:{port}"
        print(f"  [IP池] Workers全封, 使用备用IP: {ip}:{port}")
        return proxy_url
    # 全无 → 等冷却
    best = min(COOLDOWN, key=COOLDOWN.get)
    wait = max(0, COOLDOWN[best] - now)
    if wait > 0:
        print(f"  [冷却] 所有Worker+IP池耗尽, 等待 {wait:.0f}s...")
        time.sleep(min(wait, 60))
    del COOLDOWN[best]
    return best

def cool_down_proxy(proxy, seconds=600):
    """封一个 Worker, 冷却 10 分钟"""
    COOLDOWN[proxy] = time.time() + seconds
    print(f"  [冷却] {proxy.split('//')[1][:25]} 冷却 {seconds}s")

# ====== 第3层: 验证码/反爬检测关键词 ======
BLOCK_SIGNALS = [
    ("captcha", "图形验证码"),
    ("validate", "验证码"),
    ("gokuProps", "WAF 挑战"),
    ("automated access", "被标记为机器人"),
    ("Type the characters", "输入字符验证"),
    ("robot check", "机器人检测"),
    ("Access Denied", "拒绝访问"),
]

def detect_block(html):
    """检测是否被拦截, 返回 (是否, 原因)"""
    if len(html) < 3000:
        return True, f"响应过短 ({len(html)}b)"
    lower = html[:5000].lower()
    for keyword, reason in BLOCK_SIGNALS:
        if keyword.lower() in lower:
            return True, reason
    return False, ""

# 图片存放目录（优先脚本同目录，其次当前目录）
IMAGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "product_images")
if not os.path.isdir(IMAGE_DIR):
    IMAGE_DIR = os.path.join(os.getcwd(), "product_images")
if not os.path.isdir(IMAGE_DIR):
    os.makedirs(IMAGE_DIR, exist_ok=True)


def fetch(url, retries=4):
    """通过 CF Worker + IP池 双层代理抓取 — 主动轮换, 永不重复"""
    from urllib.parse import quote
    encoded = quote(url, safe='')
    cookie = ""

    for attempt in range(retries):
        # 主动轮换: 每3次请求中有1次走IP池直连, 2次走Worker
        use_ip_pool = (attempt % 3 == 1) and (_IP_POOL is not None and len(_IP_POOL) > 0)
        headers = random_headers()

        if use_ip_pool:
            # IP池直连模式: 用开放代理直接请求 (不经过Worker)
            ip_proxy = _get_next_ip()
            if ip_proxy:
                try:
                    with httpx.Client(verify=False, timeout=30, follow_redirects=True,
                                      headers=headers, proxy=ip_proxy) as client:
                        resp = client.get(url)
                        html = resp.text
                        if resp.status_code == 200 and len(html) > 1500:
                            blocked, _ = detect_block(html)
                            if not blocked:
                                print(f"  [IP池✅] {ip_proxy.split('//')[1]}")
                                return html
                except:
                    pass
            # IP池失败 → 回退到Worker
            pass

        # Worker 代理模式
        proxy = get_active_proxy(attempt)
        proxy_url = f"{proxy}/?url={encoded}"
        if cookie:
            proxy_url += f"&cookie={quote(cookie, safe='')}"

        try:
            with httpx.Client(verify=False, timeout=30, follow_redirects=True, headers=headers) as client:
                resp = client.get(proxy_url)
                html = resp.text
                proxy_cookie = resp.headers.get("x-proxy-cookie", "")
                if proxy_cookie: cookie = proxy_cookie

                if resp.status_code != 200 or len(html) < 1500:
                    print(f"  [{proxy.split('//')[1][:20]}] 异常状态={resp.status_code} len={len(html)}")
                    cool_down_proxy(proxy, 300)
                    time.sleep(3 + attempt * 2)
                    continue

                blocked, reason = detect_block(html)
                if blocked:
                    print(f"  [{proxy.split('//')[1][:20]}] {reason}, 切换...")
                    cool_down_proxy(proxy, 600)
                    time.sleep(5 + attempt * 3)
                    continue

                bm_match = re.search(r"URL=\s*['\"]?([^'\"\s>]+)", html)
                if bm_match and "bm-verify" in html[:2000]:
                    redirect = bm_match.group(1).replace("&amp;", "&")
                    if redirect.startswith("/"):
                        from urllib.parse import urlparse
                        parsed = urlparse(url)
                        redirect = f"{parsed.scheme}://{parsed.netloc}{redirect}"
                    elif not redirect.startswith("http"):
                        redirect = url.rstrip("/") + "/" + redirect.lstrip("/")
                    encoded = quote(redirect, safe='')
                    time.sleep(2)
                    continue

                return html

        except Exception as e:
            print(f"  [{proxy.split('//')[1][:20]}] 连接失败: {type(e).__name__}")
            cool_down_proxy(proxy, 120)
            time.sleep(2 + attempt)
    return None


def fetch_binary(url, retries=3):
    """通过代理池下载二进制文件（图片），随机 UA + 冷却"""
    from urllib.parse import quote
    encoded = quote(url, safe='')
    for attempt in range(retries):
        proxy = get_active_proxy(attempt)
        headers = random_headers()
        proxy_url = f"{proxy}/?url={encoded}"
        try:
            with httpx.Client(verify=False, timeout=30, follow_redirects=True, headers=headers) as client:
                resp = client.get(proxy_url)
                if resp.status_code == 200 and len(resp.content) > 5000:
                    return resp.content
                cool_down_proxy(proxy, 300)
                time.sleep(2)
        except Exception as e:
            cool_down_proxy(proxy, 120)
            time.sleep(2)
    return None


# fetch_binary moved above (unified with anti-detection layers)


def clean_text(text):
    """去除 HTML 标签和多余空白"""
    text = re.sub(r"<[^>]+>", "", text)
    return " ".join(text.split())


def extract_hi_res_images(html, asin):
    """从商品详情页提取高清图片URL（Amazon 存储完整分辨率图）"""
    images = {"main": "", "variants": [], "all": []}

    # 方法1: 从 JS 变量中提取 hi-res 图片 (最高质量)
    # Amazon 把高清图存在 'large', 'hiRes' 等 JSON 数据里
    hi_res_matches = re.findall(
        r'"hiRes"\s*:\s*"?(https?://[^"\s,]+\.(?:jpg|png))"',
        html
    )
    if not hi_res_matches:
        hi_res_matches = re.findall(
            r'"large"\s*:\s*"?(https?://[^"\s,]+\.(?:jpg|png))"',
            html
        )

    # 方法2: 从 img 标签提取 (fallback)
    img_matches = re.findall(
        r'<img[^>]*src="(https://m\.media-amazon\.com/images/I/[^"]+)"',
        html
    )

    # 方法3: 从 data-old-hires 属性
    hires_attr = re.findall(
        r'data-old-hires="(https://[^"]+\.(?:jpg|png))"',
        html
    )

    all_urls = hi_res_matches + hires_attr + img_matches

    # 去重 + 排序: 主图在前
    seen = set()
    unique_urls = []
    for u in all_urls:
        u = re.sub(r'\._[A-Z0-9_,]+_\.', '.', u)  # 去掉尺寸后缀得原图
        if u not in seen and "sprite" not in u.lower() and "transparent" not in u.lower():
            seen.add(u)
            unique_urls.append(u)

    images["main"] = unique_urls[0] if unique_urls else ""
    images["all"] = unique_urls

    return images


def download_images_for_product(asin, product_info=None):
    """下载一个商品的所有图片 — 必须抓详情页拿高清原图"""
    if product_info is None:
        product_info = {}

    os.makedirs(IMAGE_DIR, exist_ok=True)

    # 先查是否已有下载
    existing = []
    if os.path.isdir(IMAGE_DIR):
        for fname in os.listdir(IMAGE_DIR):
            if fname.startswith(asin):
                existing.append(os.path.join(IMAGE_DIR, fname))

    # 抓取详情页获取高清原图 + 提取标题/品牌等
    print(f"    抓详情页: {asin}")
    url = f"https://www.amazon.com/dp/{asin}"
    html = fetch(url)
    if not html or len(html) < 8000:
        # 被限流或 captcha —— 跳过详情页，用搜索页已有数据
        print(f"    详情页被限流(len={len(html) if html else 0})，只用搜索数据")
        return existing or []

    # 从详情页补全产品信息
    detail = parse_product_page(html, asin)
    if detail and product_info:
        for key in ["title", "brand", "bsr", "dimensions_cm", "weight", "monthly_bought"]:
            old_val = product_info.get(key, "")
            new_val = detail.get(key, "")
            is_bad = not old_val or str(old_val).startswith(("4.", "3.", "2.", "1.", "Add to", "Rated ", "Results", "Sponsored", "From the Author"))
            if new_val and is_bad:
                product_info[key] = new_val

    # 已下载过就直接返回（但信息已补全）
    if existing:
        return existing

    images = extract_hi_res_images(html, asin)
    image_urls = images["all"]
    if product_info and images["main"]:
        product_info["image_main"] = images["main"]
        product_info["images"] = images["all"]

    # 如果详情页没提取到，用搜索页的缩略图 URL 转高清
    if not image_urls and product_info.get("image_url"):
        thumb = product_info["image_url"]
        # 去掉 Amazon 的尺寸后缀得到原图
        hi_url = re.sub(r'\._[A-Z0-9_,]+_\.', '..', thumb)
        hi_url = re.sub(r'\._[A-Z0-9_,]+_\.', '.', hi_url)
        image_urls = [hi_url]

    downloaded = []
    for i, img_url in enumerate(image_urls[:7]):
        try:
            suffix = f"_{i+1}" if i > 0 else ""
            ext = ".jpg"
            if ".png" in img_url.lower():
                ext = ".png"

            filename = f"{asin}{suffix}{ext}"
            filepath = os.path.join(IMAGE_DIR, filename)

            if os.path.exists(filepath):
                downloaded.append(filepath)
                continue

            print(f"    下载图 {i+1}: {img_url[:80]}...")
            data = fetch_binary(img_url)

            if data and len(data) > 5000:
                with open(filepath, "wb") as f:
                    f.write(data)
                size_kb = len(data) // 1024
                print(f"    已保存: {filename} ({size_kb}KB)")
                downloaded.append(filepath)
            else:
                print(f"    下载失败或图片太小，跳过...")

        except Exception as e:
            print(f"    图片下载异常: {e}，跳过...")

    return downloaded


def parse_search_page(html):
    """从搜索结果页提取商品列表 — 多策略自适应"""
    products = []

    # 策略1: 按 data-asin 切分，每个 block 是一个商品卡片
    parts = re.split(r'data-asin="([A-Z0-9]{10})"', html)[1:]

    for i in range(0, len(parts), 2):
        asin = parts[i]
        block = parts[i + 1] if i + 1 < len(parts) else ""
        if not asin or len(asin) != 10:
            continue

        # --- 标题 (多策略) ---
        title = ""
        # 尝试1: h2 内的 span
        h2_m = re.search(r"<h2[^>]*>(.*?)</h2>", block, re.DOTALL)
        if h2_m:
            title = clean_text(h2_m.group(1))
        # 尝试2: aria-label on h2 or a tag
        if not title:
            aria_m = re.search(r'aria-label="([^"]{10,200})"', block)
            if aria_m:
                title = aria_m.group(1).strip()
        # 尝试3: 取最长的纯文本段(商品标题通常在20-200字符)
        if not title:
            texts = re.findall(r'>([^<]{20,200})<', block)
            candidates = [t.strip() for t in texts
                         if t.strip() and not any(kw in t.lower()
                         for kw in ['function', 'window.', 'script', 'style', 'Sponsored', 'var '])]
            if candidates:
                title = max(candidates, key=len)

        # --- 价格 ---
        price = None
        price_m = re.search(r'\$(\d+\.?\d{0,2})', block)
        if price_m:
            price = float(price_m.group(1))

        # --- 评分 ---
        rating = None
        rating_m = re.search(r'(\d\.\d)\s*out of 5', block)
        if rating_m:
            rating = float(rating_m.group(1))
        if not rating:
            aria_rating = re.search(r'aria-label="([\d.]+)\s*out of 5', block)
            if aria_rating:
                try: rating = float(aria_rating.group(1))
                except: pass

        # --- 评论数 ---
        reviews = None
        rev_m = re.search(r'<span[^>]*aria-label="(\d[\d,]*)\s*(?:ratings?|reviews?)"', block, re.IGNORECASE)
        if not rev_m:
            rev_m = re.search(r'(\d[\d,]*)\s*(?:ratings?|reviews?)', block)
        if rev_m:
            reviews = int(rev_m.group(1).replace(",", ""))
        if not reviews:
            # 尝试纯数字在评分后面
            rev_fallback = re.search(r'out of 5[^0-9]*(\d[\d,]*)', block)
            if rev_fallback:
                reviews = int(rev_fallback.group(1).replace(",", ""))

        # --- 图片 ---
        image_url = ""
        img_m = re.search(r'src="(https://m\.media-amazon\.com/images/I/[^"]+)"', block)
        if img_m:
            image_url = img_m.group(1)
        if not image_url:
            img_m2 = re.search(r'src="([^"]*\.(?:jpg|png)[^"]*)"', block)
            if img_m2:
                image_url = img_m2.group(1)

        # --- 赞助 ---
        is_sponsored = "sponsored" in block.lower() or '"s-sponsored-list"' in block.lower()

        # --- 去重: 同一 ASIN 多卡片取信息最完整的 ---
        existing = next((p for p in products if p["asin"] == asin), None)
        if existing:
            # 合并: 用更长的标题, 保留价格/评分
            if len(title) > len(existing["title"]):
                existing["title"] = title
            if price and not existing["price_usd"]:
                existing["price_usd"] = price
            if rating and not existing["rating"]:
                existing["rating"] = rating
            if reviews and not existing["review_count"]:
                existing["review_count"] = reviews
            if image_url and not existing["image_url"]:
                existing["image_url"] = image_url
        else:
            hs_code, hs_name, hs_tariff = _hs_guess(title, brand)
            products.append({
                "asin": asin,
                "title": title[:200],
                "price_usd": price,
                "rating": rating,
                "review_count": reviews,
                "image_url": image_url,
                "is_sponsored": is_sponsored,
                "hs_code": hs_code,
                "hs_name": hs_name,
                "hs_tariff": hs_tariff,
            })

    # Post-processing: filter garbage titles from search cards
    BAD = {"add to cart", "more results", "sponsored", "check each product page",
           "", "best seller", "amazon's choice", "amazon's choice: overall pick"}
    for p in products:
        t = p["title"].lower().strip().rstrip(".")
        if t in BAD or t.startswith("rated "):
            p["title"] = ""

    return products


def parse_product_page(html, asin=""):
    """从商品详情页提取详细信息"""
    detail = {"asin": asin}

    title_m = re.search(r'<span[^>]*id="productTitle"[^>]*>(.*?)</span>', html, re.DOTALL)
    detail["title"] = clean_text(title_m.group(1)) if title_m else ""

    price_m = re.search(r"priceblock_ourprice[^>]*>\$(\d+\.?\d{0,2})", html)
    if not price_m:
        price_m = re.search(r"a-price-whole[^>]*>(\d+)<", html)
    if not price_m:
        price_m = re.search(r'\$(\d+\.?\d{0,2})', html[:5000])
    detail["price_usd"] = float(price_m.group(1)) if price_m else None

    brand_m = re.search(r"Brand</span>.*?<span[^>]*>([^<]+)<", html, re.DOTALL)
    detail["brand"] = brand_m.group(1).strip() if brand_m else ""

    rating_m = re.search(r"(\d\.\d)\s*out of 5", html)
    detail["rating"] = float(rating_m.group(1)) if rating_m else None

    reviews_m = re.search(r"(\d[\d,]*)\s*(?:global\s*)?ratings?", html, re.IGNORECASE)
    detail["review_count"] = int(reviews_m.group(1).replace(",", "")) if reviews_m else None

    bsr_section = re.findall(r"Best Sellers Rank[^#]*#([\d,]+)\s*(?:in\s*([^<\n]+))?", html, re.DOTALL)
    detail["bsr"] = [{"rank": r, "category": c.strip()} for r, c in bsr_section]

    dims_match = re.search(r'product-dimensions[^>]*>.*?(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*x\s*(\d+\.?\d*)', html, re.DOTALL)
    if dims_match:
        detail["dimensions_cm"] = {
            "length": float(dims_match.group(1)),
            "width": float(dims_match.group(2)),
            "height": float(dims_match.group(3)),
        }

    weight_m = re.search(r"Item Weight[^>]*>\s*(\d+\.?\d*)\s*(pounds|ounces|kg)", html)
    if weight_m:
        detail["weight"] = f"{weight_m.group(1)} {weight_m.group(2)}"

    bought_m = re.search(r"(\d+[Kk]?\+?)\s*bought in past month", html)
    detail["monthly_bought"] = bought_m.group(1) if bought_m else ""

    # 图片
    images = extract_hi_res_images(html, asin)
    detail["image_main"] = images["main"]
    detail["images"] = images["all"]

    return detail


def search(keyword, pages=1, download_images=False):
    """搜索商品"""
    all_products = []
    products = []  # ensure defined even if all pages fail
    for page in range(1, pages + 1):
        url = f"https://www.amazon.com/s?k={quote(keyword)}&page={page}"
        print(f"Search page {page}: {url}")
        html = fetch(url)
        if html:
            products = parse_search_page(html)
            all_products.extend(products)
            print(f"  Found {len(products)} products")
        else:
            print(f"  Failed to fetch page {page}")

    if download_images and all_products:
        # 每15个产品之间额外延迟避免限流
        print(f"  Enriching from detail pages ({len(all_products)} products, slow mode)...")
        detail_count = 0
        for idx, p in enumerate(all_products):
            if detail_count >= 20:  # 一次最多补全20个，避免被限
                print(f"    ({len(all_products)-20} remaining products use search data only)")
                break
            print(f"    [{idx+1}/{len(all_products)}] ", end="")
            imgs = download_images_for_product(p["asin"], p)
            if imgs:
                detail_count += 1
            # 3-5秒间隔
            time.sleep(3 + (idx % 3))
    return all_products


def bestsellers(count=50, download_images=False):
    """抓取 Best Sellers 榜"""
    print("Fetching Best Sellers...")
    html = fetch("https://www.amazon.com/Best-Sellers/zgbs")
    if html:
        products = parse_search_page(html)
        print(f"  Found {len(products)} products")

        if download_images:
            print(f"  Downloading images...")
            for p in products[:count]:
                download_images_for_product(p["asin"], p)

        return products[:count]
    return []


def get_product(asin, download_images=False):
    """抓取单个商品详情"""
    print(f"Fetching product: {asin}")
    url = f"https://www.amazon.com/dp/{asin}"
    html = fetch(url)
    if html:
        detail = parse_product_page(html, asin)
        if download_images:
            download_images_for_product(asin, detail)
        return detail
    return None


def print_product(p, show_images=False):
    """格式化打印商品信息"""
    print(f"\n  {'='*60}")
    print(f"  ASIN:    {p.get('asin', 'N/A')}")
    print(f"  Title:   {p.get('title', 'N/A')[:100]}")
    print(f"  Brand:   {p.get('brand', 'N/A')}")
    print(f"  Price:   ${p.get('price_usd', 'N/A')}")
    print(f"  Rating:  {p.get('rating', 'N/A')} ({p.get('review_count', 'N/A')} reviews)")
    if p.get('bsr'):
        for b in p['bsr'][:3]:
            print(f"  BSR:     #{b['rank']} in {b['category']}")
    print(f"  Monthly: {p.get('monthly_bought', 'N/A')} bought")
    print(f"  Weight:  {p.get('weight', 'N/A')}")
    dims = p.get('dimensions_cm', {})
    if dims:
        print(f"  Dims:    {dims.get('length','?')}x{dims.get('width','?')}x{dims.get('height','?')} cm")

    if show_images:
        images = p.get('images', []) or []
        if p.get('image_main'):
            images = [p['image_main']] + images
        print(f"  Images:  {len(images)} pics")
        for img in images[:5]:
            print(f"           {img}")


def export_excel_111(products, filepath="选品补货表.xlsx", sheet_name="选品", embed_images=True):
    """
    按 111.xlsx 模板格式导出 — 完全匹配原始 26 列结构

    列结构:
      A=历史销量(日期)    B=总库存          C=FBA库存      D=在途库存
      E=FBA+在途+安全     F=计划补货数量     G=到货时间     H=1688下单
      I=付款时间          J=预计上架时间     K=总耗时      L=ASIN
      M=产品图片          N=30天销量        O=7天销量     P=日均销量(=O/7)
      Q=安全库存(=P*21)   R=补货数量(=Q-C-D) S=FBA配送费   T=FBA仓储费
      U=主要竞品          V=实际补货数量     W/X/Y=备注    Z=缺货原因
    """
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ========== 样式定义 ==========
    header_font = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ========== Row 1-2: 双行表头 ==========
    # Row 1 (main headers)
    h1 = [
        ("历史销量\n日期范围", 12), ("总库存\n(FBA+在途)", 10), ("FBA库存", 8), ("在途库存", 8),
        ("FBA+在途\n+安全库存", 10), ("计划补货\n数量", 8), ("到货时间", 10),
        ("1688下单", 10), ("付款时间", 10), ("预计到货\n&上架时间", 10), ("总耗时\n(采购+运输+上架)", 10),
        ("ASIN", 14), ("产品图片", 14),
        ("30天\n销量", 7), ("7天\n销量", 7), ("平均日销\n(7天日均)", 8),
        ("安全库存\n(日销×21)", 8), ("补货数量\n(安全-FBA-在途)", 10),
        ("FBA\n配送费", 7), ("FBA\n仓储费", 7), ("主要竞品\nASIN", 12),
        ("实际补货\n数量", 8), ("备注", 10), ("", 8), ("", 8), ("缺货原因", 10),
    ]

    # Row 2 (sub-header / instruction row)
    h2_instruction = "按20天FBA库存预警线: 日均销量×21天安全库存计算, FBA+在途<安全库存时触发补货。1688采购按5天, 头程上架按10天计。"

    for col_idx, (title, width) in enumerate(h1, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 2: merge all and put instruction
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=26)
    inst_cell = ws.cell(row=2, column=1, value=h2_instruction)
    inst_cell.font = Font(name="微软雅黑", size=8, italic=True, color="666666")
    inst_cell.alignment = Alignment(vertical="center")
    inst_cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # Apply border to rest of row 2 merged cells
    for c in range(2, 27):
        ws.cell(row=2, column=c).border = thin_border

    # ========== 数据行 (从 Row 3 开始) ==========
    today_str = time.strftime("%Y.%m.%d")

    for i, p in enumerate(products):
        row = i + 3
        asin = p.get("asin", "")
        ws.row_dimensions[row].height = 90  # 留够图片高度

        # --- Column A: 日期范围 (本周) ---
        ws.cell(row=row, column=1, value=today_str).alignment = center_align

        # --- Column B/C/D: 库存 (留给卖家填) ---
        # B = C + D 公式
        ws.cell(row=row, column=2).value = f"=C{row}+D{row}"
        ws.cell(row=row, column=2).alignment = center_align

        # --- Column E: FBA+在途+安全 = B + Q ---
        ws.cell(row=row, column=5).value = f"=B{row}+Q{row}"
        ws.cell(row=row, column=5).alignment = center_align

        # --- Column N: 30天销量 (采集到的评论数作为销量参考) ---
        est_30d = p.get("review_count", 0) or 0
        ws.cell(row=row, column=14, value=est_30d).alignment = center_align

        # --- Column O: 7天销量 (30天/4.3) ---
        ws.cell(row=row, column=15).value = f"=ROUND(N{row}/4.3,0)"
        ws.cell(row=row, column=15).alignment = center_align

        # --- Column P: 日均销量 = O/7 ---
        ws.cell(row=row, column=16).value = f"=ROUND(O{row}/7,1)"
        ws.cell(row=row, column=16).alignment = center_align
        ws.cell(row=row, column=16).number_format = '0.0'

        # --- Column Q: 安全库存 = P * 21 ---
        ws.cell(row=row, column=17).value = f"=ROUND(P{row}*21,0)"
        ws.cell(row=row, column=17).alignment = center_align

        # --- Column R: 补货数量 = Q - C - D ---
        ws.cell(row=row, column=18).value = f"=Q{row}-C{row}-D{row}"
        ws.cell(row=row, column=18).alignment = center_align

        # --- Column L: ASIN ---
        ws.cell(row=row, column=12, value=asin).alignment = center_align

        # --- 产品信息填入备注列 W/X ---
        info_parts = []
        if p.get("title"):
            info_parts.append(p["title"][:100])
        if p.get("brand"):
            info_parts.append(f"品牌: {p['brand']}")
        if p.get("price_usd"):
            info_parts.append(f"售价: ${p['price_usd']}")
        if p.get("rating"):
            info_parts.append(f"评分: {p['rating']} ({p.get('review_count', '')}评)")

        bsr_list = p.get("bsr") or []
        if bsr_list:
            for b in bsr_list[:2]:
                if b.get("rank"):
                    info_parts.append(f"BSR: #{b['rank']} in {b.get('category','')}")

        if p.get("monthly_bought"):
            info_parts.append(f"月销: {p['monthly_bought']}")

        if p.get("is_sponsored"):
            info_parts.append("[广告]")

        # 图片链接写入备注（方便查看）
        if p.get("image_main"):
            info_parts.append(f"图片: {p['image_main']}")
        elif p.get("image_url"):
            info_parts.append(f"图片: {p['image_url']}")

        ws.cell(row=row, column=23, value="\n".join(info_parts)).alignment = data_align

        # --- 竞品列 U (从搜索页可能有的关联商品) ---
        related = p.get("related_asins", [])
        if related:
            ws.cell(row=row, column=21, value=", ".join(related[:5])).alignment = data_align

        # --- 应用边框 ---
        for c in range(1, 27):
            ws.cell(row=row, column=c).border = thin_border

        # --- Column M: 嵌入产品图片 (中文路径 workaround) ---
        if embed_images and os.path.isdir(IMAGE_DIR):
            local_img = None
            for fname in os.listdir(IMAGE_DIR):
                if fname.startswith(asin):
                    local_img = os.path.join(IMAGE_DIR, fname)
                    break
            if local_img and os.path.exists(local_img):
                try:
                    # PIL 处理中文路径有 bug，用字节流方式加载
                    from io import BytesIO
                    with open(local_img, "rb") as f:
                        img_data = f.read()
                    img = XLImage(BytesIO(img_data))
                    img.width = 100
                    img.height = 100
                    ws.row_dimensions[row].height = 85
                    ws.add_image(img, f"M{row}")
                except Exception as e:
                    print(f"    [warn] 图片嵌入失败 {asin}: {e}")

    # ========== 汇总行 ==========
    summary_row = len(products) + 3
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=11)
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=1).alignment = center_align
    ws.cell(row=summary_row, column=14, value=f"=SUM(N3:N{summary_row-1})").alignment = center_align
    ws.cell(row=summary_row, column=15, value=f"=SUM(O3:O{summary_row-1})").alignment = center_align
    ws.cell(row=summary_row, column=18, value=f"=SUM(R3:R{summary_row-1})").alignment = center_align
    for c in range(1, 27):
        ws.cell(row=summary_row, column=c).border = thin_border

    # ========== 冻结窗格 ==========
    ws.freeze_panes = "A3"

    # ========== 打印设置 ==========
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    print(f"\n  准备保存... (Sheet1: {sheet_name} | Sheet2: 产品详情 | 商品数: {len(products)})")
    print(f"  Sheet1公式: 日均销量=O/7, 安全库存=P×21, 补货=Q-C-D")
    print(f"  库存列(C/D)留空, 填入数字后自动计算补货量")

    # ========== Sheet 2: 产品详情 (价格/品牌/重量/尺寸独立列) ==========
    ws2 = wb.create_sheet("产品详情")
    detail_headers = [
        "图片", "ASIN", "商品名称", "品牌", "售价(USD)",
        "评分", "评论数", "BSR排名", "月销量",
        "重量", "尺寸(长x宽x高 cm)", "商品链接", "是否广告"
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 列宽
    col_widths = [14, 14, 35, 14, 12, 8, 10, 28, 10, 12, 20, 40, 10]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for row_idx, p in enumerate(products, 2):
        bsr_str = "; ".join(
            f"#{b['rank']}" + (f"({b['category']})" if b.get('category') else "")
            for b in (p.get("bsr") or [])
            if b.get("rank")
        )
        dims = p.get("dimensions_cm", {})
        dims_str = f"{dims['length']}x{dims['width']}x{dims['height']}" if dims else ""

        ws2.append([
            "",  # 图片列
            p.get("asin", ""),
            p.get("title", ""),
            p.get("brand", "").replace("From the Author", ""),
            p.get("price_usd", ""),
            p.get("rating", ""),
            p.get("review_count", ""),
            bsr_str,
            p.get("monthly_bought", ""),
            p.get("weight", ""),
            dims_str,
            f"https://www.amazon.com/dp/{p.get('asin','')}",
            "是" if p.get("is_sponsored") else "否",
        ])

        ws2.row_dimensions[row_idx].height = 85

        # 嵌入图片
        if embed_images and os.path.isdir(IMAGE_DIR):
            for fname in os.listdir(IMAGE_DIR):
                if fname.startswith(p.get("asin", "")):
                    local_img = os.path.join(IMAGE_DIR, fname)
                    if os.path.exists(local_img):
                        try:
                            from io import BytesIO
                            with open(local_img, "rb") as f:
                                img_data = f.read()
                            img = XLImage(BytesIO(img_data))
                            img.width = 100; img.height = 100
                            ws2.add_image(img, f"A{row_idx}")
                        except: pass
                    break

        for c in range(1, len(detail_headers) + 1):
            ws2.cell(row=row_idx, column=c).border = thin_border

    ws2.freeze_panes = "A2"

    wb.save(filepath)
    print(f"\n已导出: {filepath}")
    print(f"  Sheet1 ({sheet_name}): 26列补货表, 公式已内置")
    print(f"  Sheet2 (产品详情): 13列, 含品牌|售价|重量|尺寸|链接")


def main():
    parser = argparse.ArgumentParser(description="Amazon 数据采集 + 图片下载工具 (输出格式=111.xlsx模板)")
    sub = parser.add_subparsers(dest="cmd")

    sub_search = sub.add_parser("search", help="关键词搜索")
    sub_search.add_argument("keyword")
    sub_search.add_argument("--pages", type=int, default=1)
    sub_search.add_argument("--images", action="store_true", help="下载商品图片")
    sub_search.add_argument("--export", type=str, help="导出Excel路径 (默认: 选品补货表.xlsx)")

    sub_product = sub.add_parser("product", help="商品详情")
    sub_product.add_argument("asin")
    sub_product.add_argument("--images", action="store_true", help="下载商品图片")

    sub_bsr = sub.add_parser("bestsellers", help="Best Sellers 榜")
    sub_bsr.add_argument("--count", type=int, default=20)
    sub_bsr.add_argument("--images", action="store_true", help="下载商品图片")
    sub_bsr.add_argument("--export", type=str, help="导出Excel路径 (默认: 选品补货表.xlsx)")

    args = parser.parse_args()

    if args.cmd == "search":
        products = search(args.keyword, args.pages, download_images=args.images)
        if args.export:
            sheet_name = args.keyword[:20].replace("/", "-")
            export_excel_111(products, args.export, sheet_name=sheet_name, embed_images=args.images)
        else:
            print(f"\n提示: 加 --export result.xlsx 导出为 111 模板格式")
        print(f"\nTotal: {len(products)} products")
        for p in products:
            print_product(p, show_images=args.images)

    elif args.cmd == "product":
        detail = get_product(args.asin, download_images=args.images)
        if detail:
            print_product(detail, show_images=args.images)

    elif args.cmd == "bestsellers":
        products = bestsellers(args.count, download_images=args.images)
        if args.export:
            export_excel_111(products, args.export, sheet_name="BestSellers", embed_images=args.images)
        else:
            print(f"\n提示: 加 --export result.xlsx 导出为 111 模板格式")
        print(f"\nTotal: {len(products)} products")
        for p in products:
            print_product(p, show_images=args.images)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
