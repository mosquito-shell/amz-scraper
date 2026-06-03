"""
物流发票自动生成器 — 选好产品后自动填 领翼-模版
用法: python logistics_generator.py 选品推荐_含评分.xlsx --country GB --site UK
"""
import json, os, sys, time, random, argparse
from datetime import datetime

# ========== FBA 地址库 (从领翼-模版提取) ==========
FBA_ADDRESSES = {
    "DE": [
        {"code":"DTM2-44145","warehouse":"DTM2","name":"DTM2-Amazon","addr1":"Kaltbandstrasse 4","city":"Dortmund","state":"North Rhine-Westphalia","zip":"44145"},
        {"code":"HAJ1-38350","warehouse":"HAJ1","name":"HAJ1-Amazon","addr1":"Zur Alten Molkerei 1","city":"Helmstedt","state":"Niedersachsen","zip":"38350"},
        {"code":"BER8-12529","warehouse":"BER8","name":"BER8-Amazon","addr1":"Am Mollenpfuhl 2","city":"Schonefeld","state":"Brandenburg","zip":"12529"},
        {"code":"BER3-14656","warehouse":"BER3","name":"BER3-Amazon","addr1":"Havellandstrasse 5","city":"Brieselang","state":"Brandenburg","zip":"14656"},
        {"code":"CGN1-56330","warehouse":"CGN1","name":"CGN1-Amazon","addr1":"Amazonstrasse 1","city":"Kobern-Gondorf","state":"Rhineland-Palatinate","zip":"56330"},
        {"code":"DUS2-47495","warehouse":"DUS2","name":"DUS2-Amazon","addr1":"Amazonstrasse 1","city":"Rheinberg","state":"Nordrhein-Westfalen","zip":"47495"},
        {"code":"FRA3-36251","warehouse":"FRA3","name":"FRA3-Amazon","addr1":"Amazonstrasse 1","city":"Bad Hersfeld","state":"Hesse","zip":"36251"},
        {"code":"HAM2-21423","warehouse":"HAM2","name":"HAM2-Amazon","addr1":"Borgwardstrasse 10","city":"Winsen an der Luhe","state":"Lower Saxony","zip":"21423"},
        {"code":"LEJ1-04347","warehouse":"LEJ1","name":"LEJ1-Amazon","addr1":"Amazonstrasse 1","city":"Leipzig","state":"Saxony","zip":"04347"},
        {"code":"STR1-75177","warehouse":"STR1","name":"STR1-Amazon","addr1":"Amazonstrasse 1","city":"Pforzheim","state":"Baden-Wurttemberg","zip":"75177"},
        {"code":"WRO1-16515","warehouse":"WRO1","name":"WRO1-Amazon","addr1":"Am Wald 1","city":"Oranienburg","state":"Brandenburg","zip":"16515"},
        {"code":"MHG9-63801","warehouse":"MHG9","name":"MHG9-Amazon","addr1":"Ossenheimer Strasse 7","city":"Kleinostheim","state":"Bavaria","zip":"63801"},
    ],
    "GB": [
        {"code":"BHX4-CV59PF","warehouse":"BHX4","name":"BHX4-Amazon","addr1":"Plot 1, Lyons Park,Lyons Dr","city":"Coventry","state":"","zip":"CV5 9PF"},
        {"code":"BHX2-LE671FB","warehouse":"BHX2","name":"BHX2-Amazon","addr1":"Beveridge Ln,","city":"Ellistown","state":"Coalville","zip":"LE67 1FB"},
        {"code":"BHX3-NN118PQ","warehouse":"BHX3","name":"BHX3-Amazon","addr1":"4 Royal Oak Way N,","city":"Daventry","state":"","zip":"NN11 8PQ"},
        {"code":"IDEAL-M297JY","warehouse":"海外仓","name":"IDEALLIFE-YH196","addr1":"Unit 2 Astley park estate, Kenndy Road","city":"MANCHESTER","state":"Tyldesley","zip":"M297JY"},
    ],
    "ES": [
        {"code":"BCN1-08820","warehouse":"BCN1","name":"BCN1-Amazon","addr1":"6-8 El Prat de Llobregat","city":"Barcelona","state":"Barcelona","zip":"08820"},
        {"code":"MAD6-45200","warehouse":"MAD6","name":"MAD6-Amazon","addr1":"AvenidaDeLaLogisticaPoligono","city":"Illescas","state":"Castile-La Mancha","zip":"45200"},
        {"code":"MAD7-45200","warehouse":"MAD7","name":"MAD7-Amazon","addr1":"Carretera antigua Madrid-Toledo 38D","city":"Illescas","state":"","zip":"45200"},
    ],
    "FR": [
        {"code":"CDG7-60300","warehouse":"CDG7","name":"CDG7-Amazon","addr1":"avenue Alain Boucher Parc","city":"Senlis","state":"Oise","zip":"60300"},
        {"code":"LIL1-59553","warehouse":"LIL1","name":"LIL1-Amazon","addr1":"1, Rue Amazon","city":"Lauwin-Planque","state":"","zip":"59553"},
        {"code":"ORY1-45770","warehouse":"ORY1","name":"ORY1-Amazon","addr1":"1401 Rue Du Champ Rouge","city":"Saran","state":"","zip":"45770"},
        {"code":"MRS1-26132","warehouse":"MRS1","name":"MRS1-Amazon","addr1":"Building II ZAC les","city":"Montelimar","state":"","zip":"26132"},
    ],
    "CZ": [
        {"code":"PRG2-25261","warehouse":"PRG2","name":"PRG2-Amazon","addr1":"K Amazonu 235","city":"Dobroviz","state":"","zip":"25261"},
    ],
}

# ========== 物流渠道 (从领翼-模版提取) ==========
SERVICE_CHANNELS = {
    "DE": [
        "欧洲空运普货德国清-UPS派",
        "欧洲空运普货德国清-GLS派",
        "欧洲卡航包税-UPS派",
        "欧洲卡航递延-UPS派",
        "欧洲铁路包税-GLS派",
    ],
    "GB": [
        "英国空运普货包税-DPD派",
        "英国空运五日提包税-DPD派",
        "英国空运普货不包税-DPD派",
        "英国卡航包税-DPD派",
        "英国海运包税-DPD派",
    ],
    "ES": [
        "欧洲空运六日提包税-GLS派",
        "欧洲海运包税经济线比利时清-GLS派",
    ],
    "FR": [
        "欧洲空运六日提包税-GLS派",
        "欧洲卡航包税-GLS派",
    ],
    "CZ": [
        "欧洲空运六日提包税-GLS派",
        "欧洲铁路包税-GLS派",
    ],
}

# ========== 海关编码 (按品类) ==========
HS_CODES = {
    "sports": "9506.99",      # 运动器材
    "badminton": "9506.59",   # 羽毛球拍
    "yoga": "6112.49",        # 运动服装
    "clothing": "6104.63",    # 女装/瑜伽裤
    "jewelry": "7117.19",     # 仿珠宝
    "bracelet": "7117.90",    # 手链
    "electronics": "8504.40", # 充电器
    "charger": "8504.40",
    "default": "9503.00",     # 默认: 玩具/杂项
}

def guess_hs_code(title, category=""):
    """根据标题和品类猜海关编码"""
    t = (title + " " + category).lower()
    for kw, code in HS_CODES.items():
        if kw in t:
            return code
    return HS_CODES["default"]

def get_default_address(country_code):
    """获取国家对应的默认 FBA 地址"""
    addrs = FBA_ADDRESSES.get(country_code, FBA_ADDRESSES.get("GB", []))
    return addrs[0] if addrs else None

def get_channel(country_code, tax_mode="包税"):
    """获取推荐物流渠道"""
    channels = SERVICE_CHANNELS.get(country_code, SERVICE_CHANNELS.get("GB", []))
    # 优先匹配包税/不包税
    for ch in channels:
        if tax_mode in ch:
            return ch
    return channels[0] if channels else "英国空运普货包税-DPD派"

def estimate_weight(dims, density=0.003):
    """根据尺寸估算重量 (kg)"""
    if isinstance(dims, dict):
        vol = dims.get("length", 30) * dims.get("width", 20) * dims.get("height", 5)
        return max(0.1, round(vol * density, 2))
    return 0.5

def generate_order_id():
    """生成 FBA 货件编号"""
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    return "FBA" + "".join(random.choices(chars, k=14))

def generate_invoice(products, country="GB", site="UK", output_path=None):
    """生成领翼格式的物流发票 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return None

    addr = get_default_address(country)
    if not addr:
        print(f"未知国家代码: {country}")
        return None

    wb = openpyxl.Workbook()
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold = Font(bold=True, size=10)

    # ====== Sheet 1: 发票 ======
    ws = wb.active
    ws.title = "物流发票"

    # 基本信息区 (左侧 A2:B20)
    basic_info = [
        ("服务*", get_channel(country, "包税")),
        ("地址库编码*", addr["code"]),
        ("收件人姓名*", f"{addr['warehouse']}-Amazon"),
        ("收件人公司", "Amazon Fulfillment Center"),
        ("收件人地址一*", addr["addr1"]),
        ("收件人城市*", addr["city"]),
        ("收件人邮编*", addr["zip"]),
        ("收件人国家代码*", country),
        ("收件人电话", "88888"),
        ("客户订单号*", generate_order_id()),
    ]

    for i, (label, value) in enumerate(basic_info):
        r = i + 1
        c = ws.cell(row=r, column=1, value=label)
        c.font = bold; c.border = thin; c.fill = header_fill
        c = ws.cell(row=r, column=2, value=value)
        c.border = thin

    # 右侧参数区 (E2:I20)
    right_params = [
        ("箱数*", len(products)),
        ("带电*", "否"),
        ("带磁*", "否"),
        ("粉末*", "否"),
        ("液体*", "否"),
        ("报关方式*", f"{site}站-买单报关"),
        ("交税方式*", "包税"),
        ("备注*", f"交货时间: 按货代排期"),
    ]
    for i, (label, value) in enumerate(right_params):
        r = i + 1
        c = ws.cell(row=r, column=5, value=label)
        c.font = bold; c.border = thin; c.fill = header_fill
        c = ws.cell(row=r, column=6, value=value)
        c.border = thin

    # 产品清单表头 (Row 23-25)
    headers = [
        ("货箱编号*", 10), ("产品英文品名*", 35), ("产品中文品名*", 15),
        ("申报数量*", 8), ("申报单位", 6), ("申报单价(USD)*", 10),
        ("重量(KG)*", 8), ("长(CM)*", 6), ("宽(CM)*", 6), ("高(CM)*", 6),
        ("海关编码*", 12), ("品牌*", 12), ("材质*", 10),
        ("用途*", 15), ("销售链接*", 35), ("图片*", 25), ("SKU", 12),
    ]

    for ci, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=23, column=ci, value=h)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 产品数据
    for ri, p in enumerate(products):
        row = 26 + ri
        asin = p.get("asin", "")
        title = p.get("title", "")[:100]
        hs = guess_hs_code(title, str(p.get("bsr", [{}])[0].get("category", "") if p.get("bsr") else ""))
        brand = p.get("brand", "").replace("From the Author", "自有品牌")
        weight = estimate_weight(p.get("dimensions_cm"))
        dims = p.get("dimensions_cm", {})
        length = dims.get("length", 30) if dims.get("length") else 30
        width = dims.get("width", 20) if dims.get("width") else 20
        height = dims.get("height", 5) if dims.get("height") else 5
        price = p.get("price_usd", 0) or 0
        qty = max(1, int(p.get("monthly_bought", "100").replace("+","").replace("K","000") or 100))

        row_data = [
            f"BOX-{ri+1}", title, title[:20], qty, "PCS", round(price*0.6, 2),
            round(weight*qty, 2), length, width, height,
            hs, brand, "塑料/金属",
            "户外运动" if "sport" in title.lower() or "badminton" in title.lower() else "日常使用",
            f"https://www.amazon.com/dp/{asin}",
            p.get("image_main", "") or p.get("image_url", "") or "",
            asin,
        ]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=(ci in [2,3,15,16]))
            if ci == 6: c.number_format = '0.00'

    # ====== Sheet 2: FBA 地址库 ======
    ws2 = wb.create_sheet("FBA地址库")
    addr_headers = ["地址编码", "FBA仓库", "联系人", "地址", "城市", "省份", "国家", "邮编"]
    for ci, h in enumerate(addr_headers, 1):
        ws2.cell(row=1, column=ci, value=h).font = bold
        ws2.column_dimensions[get_column_letter(ci)].width = 18

    row2 = 2
    for cntry, addrs in FBA_ADDRESSES.items():
        for a in addrs:
            ws2.cell(row=row2, column=1, value=a["code"])
            ws2.cell(row=row2, column=2, value=a["warehouse"])
            ws2.cell(row=row2, column=3, value=a["name"])
            ws2.cell(row=row2, column=4, value=a["addr1"])
            ws2.cell(row=row2, column=5, value=a["city"])
            ws2.cell(row=row2, column=6, value=a.get("state",""))
            ws2.cell(row=row2, column=7, value=cntry)
            ws2.cell(row=row2, column=8, value=a["zip"])
            row2 += 1

    # ====== Sheet 3: 物流渠道 ======
    ws3 = wb.create_sheet("物流渠道")
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 40
    ws3.cell(row=1, column=1, value="国家").font = bold
    ws3.cell(row=1, column=2, value="可选渠道").font = bold
    row3 = 2
    for cntry, chs in SERVICE_CHANNELS.items():
        for ch in chs:
            ws3.cell(row=row3, column=1, value=cntry)
            ws3.cell(row=row3, column=2, value=ch)
            row3 += 1

    # 保存
    if not output_path:
        output_path = f"物流发票_{country}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(output_path)
    print(f"\n✅ 物流发票已生成: {output_path}")
    print(f"   国家: {country} | FBA仓库: {addr['warehouse']}-{addr['city']}")
    print(f"   渠道: {get_channel(country)}")
    print(f"   产品数: {len(products)} 箱 | 申报总重: {sum(estimate_weight(p.get('dimensions_cm'))*max(1,int(str(p.get('monthly_bought','100')).replace('+','').replace('K','000') or 100)) for p in products):.1f} kg")
    return output_path

# ========== 从选品 Excel 读取产品 ==========
def load_products_from_excel(filepath):
    """从选品推荐 Excel 读取已选产品"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = None
        for sn in wb.sheetnames:
            if '推荐' in sn or '选品' in sn or '排名' in sn:
                ws = wb[sn]; break
        if not ws:
            ws = wb.active

        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[3] or str(row[3]).strip() == '':
                continue
            asin = str(row[3]).strip()  # Column D = ASIN
            if len(asin) != 10 or not asin.startswith('B'):
                continue
            # 只取标记为 "Y" 或 "✅" 的行
            keep_col = str(row[0]) if row[0] else ''
            if 'N' in keep_col or '❌' in keep_col:
                continue

            p = {"asin": asin}
            if len(row) > 4: p["title"] = str(row[4]) if row[4] else ""
            if len(row) > 5: p["brand"] = str(row[5]) if row[5] else ""
            if len(row) > 6:
                try: p["price_usd"] = float(row[6])
                except: pass
            products.append(p)
        return products
    except Exception as e:
        print(f"读取Excel失败: {e}")
        return []

def main():
    parser = argparse.ArgumentParser(description="物流发票生成器")
    parser.add_argument("file", nargs="?", help="选品Excel路径")
    parser.add_argument("--country", default="GB", choices=["GB","DE","ES","FR","CZ"], help="目标国家 (默认GB)")
    parser.add_argument("--site", default="UK", help="站点 (UK/DE/ES/FR/CZ)")
    parser.add_argument("--output", help="输出路径")
    parser.add_argument("--demo", action="store_true", help="生成示例发票")
    args = parser.parse_args()

    if args.demo or not args.file:
        # 示例产品
        demo = [
            {"asin":"B08BNJT83C","title":"HIRALIY Badminton Rackets Set - 4 Rackets, 12 Birdies","brand":"HIRALIY","price_usd":39.99,"dimensions_cm":{"length":38,"width":25,"height":8},"monthly_bought":"200","image_main":"","bsr":[{"category":"Sports & Outdoors"}]},
            {"asin":"B083PB9GFY","title":"CRZ YOGA Womens Butterluxe Yoga Leggings","brand":"CRZ YOGA","price_usd":28.00,"dimensions_cm":{"length":25,"width":18,"height":3},"monthly_bought":"1000","image_main":"","bsr":[{"category":"Clothing, Shoes & Jewelry"}]},
            {"asin":"B0DX6Q2K5D","title":"Gokeey Gold Bracelets Set Women 14K Gold Plated","brand":"Gokeey","price_usd":12.56,"dimensions_cm":{"length":10,"width":8,"height":2},"monthly_bought":"200","image_main":"","bsr":[{"category":"Clothing, Shoes & Jewelry"}]},
        ]
        generate_invoice(demo, args.country, args.site, args.output)
    else:
        products = load_products_from_excel(args.file)
        if not products:
            print("未找到产品，使用 --demo 生成示例")
            return
        generate_invoice(products, args.country, args.site, args.output)

if __name__ == "__main__":
    main()
