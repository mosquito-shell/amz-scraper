"""
批量采集脚本 — 一键跑完所有关键词
用法: python batch_crawl.py
配置文件: keywords.txt (每行一个关键词)
输出: 批量采集_YYYYMMDD.xlsx
"""
import os
import sys
import time
import re
from datetime import datetime
from urllib.parse import quote

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from scraper import fetch, parse_search_page, parse_product_page
from scorer import score_all, get_recommendation

urllib3_import = __import__('urllib3')
urllib3_import.disable_warnings()

KEYWORDS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "keywords.txt")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ====== 默认关键词（如果没有 keywords.txt）======
DEFAULT_KEYWORDS = [
    "badminton set backyard",
    "yoga pants women high waist",
    "bracelet women gold stackable",
    "portable bluetooth speaker",
    "yoga mat non slip",
    "resistance bands set",
    "pilates socks grip",
    "stainless steel water bottle",
    "led strip lights bedroom",
    "phone stand adjustable desk",
    "cable management desk",
    "silicone baking mat",
    "measuring spoons set",
    "pet grooming gloves",
    "baby sensory toys",
]


def load_keywords():
    """加载关键词列表"""
    if os.path.exists(KEYWORDS_FILE):
        with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f.readlines() if l.strip() and not l.startswith("#")]
        if lines:
            print(f"从 {KEYWORDS_FILE} 加载 {len(lines)} 个关键词")
            return lines

    # 创建默认文件
    print(f"未找到 {KEYWORDS_FILE}，已创建默认关键词文件")
    with open(KEYWORDS_FILE, "w", encoding="utf-8") as f:
        f.write("# 亚马逊选品批量采集关键词\n")
        f.write("# 每行一个关键词，以 # 开头的行为注释\n\n")
        for kw in DEFAULT_KEYWORDS:
            f.write(kw + "\n")
    return DEFAULT_KEYWORDS


def crawl_keyword(keyword, max_pages=1, sleep_between=8):
    """采集单个关键词的搜索结果"""
    print(f"\n{'='*60}")
    print(f"  搜索: {keyword}")
    print(f"{'='*60}")

    all_products = []
    for page in range(1, max_pages + 1):
        url = f"https://www.amazon.com/s?k={quote(keyword)}&page={page}"
        print(f"  第 {page} 页...", end=" ")
        html = fetch(url)
        if not html:
            print("失败（可能被拦截）")
            break

        products = parse_search_page(html)
        # 标记来源关键词
        for p in products:
            p["_keyword"] = keyword
            p["_page"] = page
        all_products.extend(products)
        print(f"抓到 {len(products)} 个商品")

        if page < max_pages:
            # 随机间隔（模拟人类）
            wait = sleep_between + (page % 3) * 2
            print(f"  等待 {wait}s...")
            time.sleep(wait)

    print(f"  共计: {len(all_products)} 个商品")
    return all_products


def deduplicate(products):
    """去重：同一 ASIN 保留信息最完整的"""
    seen = {}
    for p in products:
        asin = p.get("asin", "")
        if not asin: continue
        if asin in seen:
            # 保留标题更长的版本
            if len(p.get("title", "")) > len(seen[asin].get("title", "")):
                p["_keywords"] = list(set(
                    (seen[asin].get("_keywords", []) or []) +
                    [seen[asin].get("_keyword", "")]
                ))
                seen[asin] = p
            seen[asin]["_keywords"] = list(set(
                (seen[asin].get("_keywords", []) or []) +
                [p.get("_keyword", "")]
            ))
        else:
            p["_keywords"] = [p.get("_keyword", "")]
            seen[asin] = p
    return list(seen.values())


def export_all(products, filepath):
    """导出完整 Excel（含打分）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    # ====== Sheet 1: 全部排名 ======
    ws = wb.active
    ws.title = "选品排名"

    headers = ["排名", "推荐分", "推荐", "ASIN", "商品名称", "品牌", "售价(USD)",
               "评分", "评论数", "BSR", "来源关键词", "价格分", "需求分", "竞争分",
               "品牌分", "安全分", "社交分", "商品链接"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    # 给每个产品补默认字段，确保 scorer 不报错
    for p in products:
        if not p.get("review_count"): p["review_count"] = 0
        if not p.get("rating"): p["rating"] = 0
        if not p.get("price_usd"): p["price_usd"] = 0
        if not p.get("brand"): p["brand"] = ""
        p["bsr"] = p.get("bsr") or [{"rank": "999999", "category": ""}]
        p["_tmStatus"] = "待查"
    ranked = score_all(products)

    for ri, p in enumerate(ranked, 2):
        d = p.get("_details", {})
        bsr = (p.get("bsr") or [{}])[0].get("rank", "") if p.get("bsr") else ""
        kw_str = ", ".join(p.get("_keywords", [p.get("_keyword", "")])[:3])
        row_data = [
            ri-1, p.get("score", 0), get_recommendation(p.get("score", 0)),
            p.get("asin", ""), p.get("title", "")[:100], p.get("brand", ""),
            p.get("price_usd", ""), p.get("rating", ""), p.get("review_count", ""),
            bsr, kw_str,
            round((d.get("price", 0) or 0) * 0.2, 1),
            round((d.get("demand", 0) or 0) * 0.25, 1),
            round((d.get("competition", 0) or 0) * 0.2, 1),
            round((d.get("brand", 0) or 0) * 0.15, 1),
            round((d.get("safety", 0) or 0) * 0.1, 1),
            round((d.get("social", 0) or 0) * 0.1, 1),
            f'https://www.amazon.com/dp/{p.get("asin","")}',
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [5, 11]))

        # 颜色
        s = p.get("score", 0)
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if s >= 70 else \
               PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if s >= 50 else \
               PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.cell(row=ri, column=2).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{len(ranked)+1}"

    # 列宽
    widths = [5, 7, 10, 14, 40, 14, 9, 5, 8, 14, 25, 6, 6, 6, 6, 6, 6, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ====== Sheet 2: 111 补货表 ======
    ws2 = wb.create_sheet("补货表(111模板)")
    h2 = ["日期", "总库存", "FBA库存", "在途库存", "FBA+安全", "计划补货", "到货",
          "1688下单", "付款时间", "预计上架", "总耗时", "ASIN", "图片",
          "30天销量", "7天销量", "日均销量", "安全库存", "补货量",
          "FBA配送", "FBA仓储", "竞品", "实际补货", "备注", "", "", "缺货"]
    for ci, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    today = time.strftime("%Y.%m.%d")
    for ri, p in enumerate(ranked, 2):
        ws2.cell(row=ri, column=1, value=today).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=2, value=f"=C{ri}+D{ri}")
        ws2.cell(row=ri, column=5, value=f"=B{ri}+Q{ri}")
        ws2.cell(row=ri, column=12, value=p.get("asin", "")).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=14, value=p.get("review_count", 0) or 0).alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=15, value=f"=ROUND(N{ri}/4.3,0)")
        ws2.cell(row=ri, column=16, value=f"=ROUND(O{ri}/7,1)")
        ws2.cell(row=ri, column=17, value=f"=ROUND(P{ri}*21,0)")
        ws2.cell(row=ri, column=18, value=f"=Q{ri}-C{ri}-D{ri}")
        note = f"[{p.get('score', 0)}分] {get_recommendation(p.get('score', 0))}\n{p.get('title', '')[:80]}"
        ws2.cell(row=ri, column=23, value=note).alignment = Alignment(vertical="center", wrap_text=True)
        for c in range(1, 27):
            ws2.cell(row=ri, column=c).border = thin
    ws2.freeze_panes = "A2"

    # ====== Sheet 3: 关键词统计 ======
    ws3 = wb.create_sheet("关键词分析")
    kw_stats = {}
    for p in products:
        for kw in (p.get("_keywords", []) or [p.get("_keyword", "")]):
            if not kw: continue
            if kw not in kw_stats: kw_stats[kw] = {"total": 0, "highscore": 0, "avg_price": 0, "prices": []}
            kw_stats[kw]["total"] += 1
            if p.get("score", 0) >= 70: kw_stats[kw]["highscore"] += 1
            if p.get("price_usd"): kw_stats[kw]["prices"].append(p["price_usd"])

    ws3.cell(row=1, column=1, value="关键词").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="商品数").font = Font(bold=True)
    ws3.cell(row=1, column=3, value="强推(>70)").font = Font(bold=True)
    ws3.cell(row=1, column=4, value="均价").font = Font(bold=True)
    ws3.cell(row=1, column=5, value="命中率").font = Font(bold=True)

    for ri, (kw, stats) in enumerate(sorted(kw_stats.items(), key=lambda x: -x[1]["total"]), 2):
        avg_p = round(sum(stats["prices"]) / len(stats["prices"]), 2) if stats["prices"] else 0
        hit_rate = round(stats["highscore"] / stats["total"] * 100) if stats["total"] else 0
        ws3.cell(row=ri, column=1, value=kw)
        ws3.cell(row=ri, column=2, value=stats["total"])
        ws3.cell(row=ri, column=3, value=stats["highscore"])
        ws3.cell(row=ri, column=4, value=avg_p)
        ws3.cell(row=ri, column=5, value=f"{hit_rate}%")

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 10

    wb.save(filepath)
    print(f"\n  已导出: {filepath}")
    return filepath


def main():
    print("=" * 60)
    print("  亚马逊批量选品采集系统")
    print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    keywords = load_keywords()
    all_products = []

    for i, kw in enumerate(keywords):
        print(f"\n[{i+1}/{len(keywords)}] {kw}")
        start = time.time()
        products = crawl_keyword(kw, max_pages=1, sleep_between=6)
        all_products.extend(products)
        elapsed = time.time() - start
        print(f"  耗时: {elapsed:.0f}s | 累计: {len(all_products)} 商品")

        # 每 3 个关键词额外休息
        if (i + 1) % 3 == 0 and i + 1 < len(keywords):
            wait = 20
            print(f"\n  休息 {wait}s 避免触发限流...")
            time.sleep(wait)

    # 去重
    print(f"\n去重前: {len(all_products)} | 去重后: ", end="")
    all_products = deduplicate(all_products)
    print(f"{len(all_products)}")

    # 打分
    print(f"打分中...")
    score_all(all_products)

    # 统计
    strong = len([p for p in all_products if p.get("score", 0) >= 70])
    avg_price = sum(p.get("price_usd", 0) or 0 for p in all_products)
    avg_price = round(avg_price / len(all_products), 2) if all_products else 0

    print(f"\n{'='*60}")
    print(f"  采集完成")
    print(f"  总商品: {len(all_products)} | 强推(>70分): {strong} | 均价: ${avg_price}")
    print(f"{'='*60}")

    # 导出
    output = os.path.join(OUTPUT_DIR, f"批量采集_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    export_all(all_products, output)

    # 自动上传到后端
    print(f"\n上传到后端...")
    try:
        import requests as _requests
        clean = []
        for p in all_products:
            clean.append({
                "asin": p.get("asin", ""),
                "title": p.get("title", "")[:200],
                "brand": (p.get("brand", "") or "").replace("Visit the ", "").replace(" Store", ""),
                "price": str(p.get("price_usd", "") or ""),
                "rating": str(p.get("rating", "") or ""),
                "reviews": p.get("review_count", 0) or 0,
                "bsr": (p.get("bsr", [{}]) or [{}])[0].get("rank", "") if isinstance(p.get("bsr"), list) else "",
                "monthly": str(p.get("monthly", "") or ""),
                "shipping": p.get("_shipping", "") or "",
                "image": p.get("image_url", "") or "",
                "score": p.get("score", 0)
            })
        resp = _requests.post("https://api.tsscjn.top/api/products",
            json={"products": clean, "source": "batch_crawl", "time": datetime.now().isoformat()},
            timeout=15)
        data = resp.json()
        print(f"  上传完成: {data.get('total', '?')} 条 (+{data.get('added', '?')} new)")
    except Exception as e:
        print(f"  上传失败: {e}")

    # 输出 Top 10
    print(f"\n  Top 10 推荐:")
    for i, p in enumerate(all_products[:10]):
        kw = ", ".join((p.get("_keywords", []) or [p.get("_keyword", "")])[:2])
        print(f"  {i+1:2d}. [{p.get('score',0):.1f}] ${p.get('price_usd','?'):>6} {p.get('title','')[:60]} [{kw}]")


if __name__ == "__main__":
    main()
